from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference


OUTPUT_FILE = "transport_business_accounting_departure_return.xlsx"
SHEET_NAME = "Daily Route Record"


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    headers = [
        "Departure Date",
        "Return Date",
        "Trip Days",
        "Route",
        "Diesel Expense",
        "Oil Expense",
        "Ration / Food Expense",
        "Mobile Expense",
        "Misc Expense",
        "Driver Salary",
        "Garage Expense",
        "Service Expense",
        "Total Expenses",
        "Income 1",
        "Income 2",
        "Income 3",
        "Income 4",
        "Total Income",
        "Additional Expense Detail",
        "Additional Expense Amount",
        "Profit",
        "Loss",
    ]

    # Header color palette
    fill_date = PatternFill("solid", fgColor="DDEBF7")     # light blue
    fill_expense = PatternFill("solid", fgColor="F8CBAD")  # light red
    fill_total = PatternFill("solid", fgColor="FFF2CC")    # yellow
    fill_income = PatternFill("solid", fgColor="E2F0D9")   # light green
    fill_additional = PatternFill("solid", fgColor="FCE4D6")  # light orange
    fill_profit = PatternFill("solid", fgColor="E4DFEC")   # light purple

    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

        # Header fills by column group
        if col in (1, 2, 3):
            cell.fill = fill_date
        elif 5 <= col <= 12:
            cell.fill = fill_expense
        elif col == 13:
            cell.fill = fill_total
        elif 14 <= col <= 17:
            cell.fill = fill_income
        elif col == 18:
            cell.fill = fill_total
        elif 19 <= col <= 20:
            cell.fill = fill_additional
        elif col == 21:
            cell.fill = PatternFill("solid", fgColor="E2F0D9")  # light green
        elif col == 22:
            cell.fill = PatternFill("solid", fgColor="F8CBAD")  # light red

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set column widths for better readability
    widths = {
        "A": 14,
        "B": 14,
        "C": 10,
        "D": 20,
        "E": 14,
        "F": 12,
        "G": 20,
        "H": 14,
        "I": 12,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 15,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 14,
        "S": 30,
        "T": 18,
        "U": 14,
        "V": 14,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    start_row = 2
    excel_max_row = 1_048_576

    # Keep date input simple: no strict popup validation.
    # Users can type dates directly (e.g., 26/2/2026).

    # Keep one starter data row and use an Excel Table so formulas auto-extend
    # as the user adds new rows (up to Excel's row limit).
    for col in range(1, 23):
        ws.cell(row=start_row, column=col).border = thin_border

    # Use standard A1 formulas in the first data row.
    # Excel converts these to calculated columns inside the table for new rows.
    ws["C2"] = (
        '=IF(AND(A2<>"",B2<>""),IFERROR('
        '(IF(ISNUMBER(B2),B2,DATE(VALUE(RIGHT(B2,LEN(B2)-FIND("/",B2,FIND("/",B2)+1))),'
        'VALUE(MID(B2,FIND("/",B2)+1,FIND("/",B2,FIND("/",B2)+1)-FIND("/",B2)-1)),'
        'VALUE(LEFT(B2,FIND("/",B2)-1)))))'
        '-'
        '(IF(ISNUMBER(A2),A2,DATE(VALUE(RIGHT(A2,LEN(A2)-FIND("/",A2,FIND("/",A2)+1))),'
        'VALUE(MID(A2,FIND("/",A2)+1,FIND("/",A2,FIND("/",A2)+1)-FIND("/",A2)-1)),'
        'VALUE(LEFT(A2,FIND("/",A2)-1)))))'
        ',""),"")'
    )
    ws["M2"] = "=SUM(E2:L2)"
    ws["R2"] = "=SUM(N2:Q2)"
    ws["U2"] = "=MAX(R2-(M2+T2),0)"
    ws["V2"] = "=MAX((M2+T2)-R2,0)"

    # Force manual date input as text to avoid locale auto-conversion issues.
    # Users should type exactly dd/mm/yyyy (e.g., 26/02/2026).
    ws["A2"].number_format = "@"
    ws["B2"].number_format = "@"
    ws["C2"].number_format = "0"
    ws["D2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["S2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    for col_letter in ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "T", "U", "V"]:
        ws[f"{col_letter}2"].number_format = "#,##0.00"

    data_table = Table(displayName="DailyRouteRecordTable", ref="A1:V2")
    data_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(data_table)

    # Keep initial date cells blank for manual entry.
    ws["A2"] = None
    ws["B2"] = None

    # Visible guidance for the expected manual format
    ws["A3"] = "Type date as dd/mm/yyyy"
    ws["B3"] = "Type date as dd/mm/yyyy"
    ws["A3"].font = Font(color="7F7F7F", italic=True)
    ws["B3"].font = Font(color="7F7F7F", italic=True)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["B3"].alignment = Alignment(horizontal="center")

    # Highlight positive profit/loss values
    ws.conditional_formatting.add(
        f"U2:U{excel_max_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"V2:V{excel_max_row}",
        FormulaRule(
            formula=["AND(ISNUMBER(V2),V2>0)"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True),
        ),
    )

    create_monthly_summary_sheet(wb)
    create_yearly_summary_sheet(wb)
    create_dashboard_sheet(wb)

    return wb


def create_monthly_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Monthly Summary")

    headers = [
        "Year",
        "Month No",
        "Month Name",
        "Trips",
        "Total Expenses",
        "Total Income",
        "Additional Expenses",
        "Profit",
        "Loss",
    ]

    fill_header = PatternFill("solid", fgColor="D9E1F2")
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_header
        cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    widths = {"A": 12, "B": 10, "C": 16, "D": 10, "E": 16, "F": 14, "G": 20, "H": 14, "I": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, 14):
        ws[f"A{row}"] = "=YEAR(TODAY())"
        ws[f"B{row}"] = row - 1
        ws[f"C{row}"] = f'=TEXT(DATE(A{row},B{row},1),"mmmm")'

        ws[f"D{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*(IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=B{row}))'
        )
        ws[f"E{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*(IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=B{row})'
            f'*N(DailyRouteRecordTable[Total Expenses]))'
        )
        ws[f"F{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*(IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=B{row})'
            f'*N(DailyRouteRecordTable[Total Income]))'
        )
        ws[f"G{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*(IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=B{row})'
            f'*N(DailyRouteRecordTable[Additional Expense Amount]))'
        )
        ws[f"H{row}"] = f"=MAX(F{row}-(E{row}+G{row}),0)"
        ws[f"I{row}"] = f"=MAX((E{row}+G{row})-F{row},0)"

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border

        ws[f"E{row}"].number_format = "#,##0.00"
        ws[f"F{row}"].number_format = "#,##0.00"
        ws[f"G{row}"].number_format = "#,##0.00"
        ws[f"H{row}"].number_format = "#,##0.00"
        ws[f"I{row}"].number_format = "#,##0.00"

    ws.conditional_formatting.add(
        "H2:H13",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        "I2:I13",
        FormulaRule(
            formula=["AND(ISNUMBER(I2),I2>0)"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True),
        ),
    )


def create_yearly_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Yearly Summary")

    headers = [
        "Year",
        "Trips",
        "Total Expenses",
        "Total Income",
        "Additional Expenses",
        "Profit",
        "Loss",
    ]

    fill_header = PatternFill("solid", fgColor="E2F0D9")
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_header
        cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    widths = {"A": 12, "B": 10, "C": 16, "D": 14, "E": 20, "F": 14, "G": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # 10-year view: current year - 4 to current year + 5
    for row in range(2, 12):
        ws[f"A{row}"] = f"=YEAR(TODAY())-4+ROW()-2"
        ws[f"B{row}"] = f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row}))'
        ws[f"C{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*N(DailyRouteRecordTable[Total Expenses]))'
        )
        ws[f"D{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*N(DailyRouteRecordTable[Total Income]))'
        )
        ws[f"E{row}"] = (
            f'=SUMPRODUCT((IFERROR(--RIGHT(DailyRouteRecordTable[Departure Date],4),0)=A{row})'
            f'*N(DailyRouteRecordTable[Additional Expense Amount]))'
        )
        ws[f"F{row}"] = f"=MAX(D{row}-(C{row}+E{row}),0)"
        ws[f"G{row}"] = f"=MAX((C{row}+E{row})-D{row},0)"

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border

        ws[f"C{row}"].number_format = "#,##0.00"
        ws[f"D{row}"].number_format = "#,##0.00"
        ws[f"E{row}"].number_format = "#,##0.00"
        ws[f"F{row}"].number_format = "#,##0.00"
        ws[f"G{row}"].number_format = "#,##0.00"

    ws.conditional_formatting.add(
        "F2:F11",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        "G2:G11",
        FormulaRule(
            formula=["AND(ISNUMBER(G2),G2>0)"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True),
        ),
    )


def create_dashboard_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")

    ws.merge_cells("A1:L1")
    ws["A1"] = "Transport Business Dashboard"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Dashboard controls
    ws["A2"] = "Theme"
    ws["B2"] = "Light"
    ws["A6"] = "Month Selector"
    ws["B6"] = "All"
    ws["C6"] = "Pick month to update route-wise charts"
    ws["A2"].font = Font(bold=True)
    ws["A6"].font = Font(bold=True)
    ws["C6"].font = Font(color="666666", italic=True)

    widths = {
        "A": 14, "B": 12, "C": 14, "D": 12, "E": 14, "F": 12,
        "G": 14, "H": 12, "I": 14, "J": 12, "K": 14, "L": 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    selector_fill = PatternFill("solid", fgColor="F2F2F2")
    ws["B2"].fill = selector_fill
    ws["B6"].fill = selector_fill
    ws["B2"].border = thin_border
    ws["B6"].border = thin_border
    ws["B2"].alignment = Alignment(horizontal="center")
    ws["B6"].alignment = Alignment(horizontal="center")

    theme_dv = DataValidation(type="list", formula1='"Light,Dark"', allow_blank=False)
    month_dv = DataValidation(
        type="list",
        formula1='"All,January,February,March,April,May,June,July,August,September,October,November,December"',
        allow_blank=False,
    )
    ws.add_data_validation(theme_dv)
    ws.add_data_validation(month_dv)
    theme_dv.add("B2")
    month_dv.add("B6")

    kpi_cards = [
        ("A:B", "Total Trips", '=SUM(\'Monthly Summary\'!D2:D13)', "D9E1F2"),
        ("C:D", "Total Income", '=SUM(\'Monthly Summary\'!F2:F13)', "E2F0D9"),
        ("E:F", "Total Expenses", '=SUM(\'Monthly Summary\'!E2:E13)', "FCE4D6"),
        ("G:H", "Total Profit", '=SUM(\'Monthly Summary\'!H2:H13)', "C6EFCE"),
        ("I:J", "Total Loss", '=SUM(\'Monthly Summary\'!I2:I13)', "FFC7CE"),
        ("K:L", "Net Result", "=G4-I4", "FFF2CC"),
    ]

    for col_range, label, formula, color in kpi_cards:
        start_col, end_col = col_range.split(":")
        ws.merge_cells(f"{start_col}3:{end_col}3")
        ws.merge_cells(f"{start_col}4:{end_col}4")

        ws[f"{start_col}3"] = label
        ws[f"{start_col}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{start_col}3"].font = Font(bold=True, color="000000")
        ws[f"{start_col}3"].fill = PatternFill("solid", fgColor=color)

        ws[f"{start_col}4"] = formula
        ws[f"{start_col}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{start_col}4"].font = Font(size=12, bold=True)
        ws[f"{start_col}4"].number_format = "#,##0.00"
        ws[f"{start_col}4"].fill = PatternFill("solid", fgColor=color)

        start_idx = ord(start_col) - 64
        end_idx = ord(end_col) - 64
        for row in range(3, 5):
            for col in range(start_idx, end_idx + 1):
                ws.cell(row=row, column=col).border = thin_border

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 24

    # Theme formatting: dark mode overlay
    dark_fill = PatternFill("solid", fgColor="1F1F1F")
    dark_font = Font(color="FFFFFF")
    ws.conditional_formatting.add(
        "A1:L2",
        FormulaRule(formula=['$B$2="Dark"'], fill=dark_fill, font=dark_font),
    )
    ws.conditional_formatting.add(
        "A5:L45",
        FormulaRule(formula=['$B$2="Dark"'], fill=dark_fill, font=dark_font),
    )

    # Monthly income vs expenses line chart
    monthly_line = LineChart()
    monthly_line.title = "Monthly Income vs Expenses"
    monthly_line.style = 2
    monthly_line.y_axis.title = "Amount"
    monthly_line.x_axis.title = "Month"
    data = Reference(wb["Monthly Summary"], min_col=5, max_col=6, min_row=1, max_row=13)
    cats = Reference(wb["Monthly Summary"], min_col=3, min_row=2, max_row=13)
    monthly_line.add_data(data, titles_from_data=True)
    monthly_line.set_categories(cats)
    monthly_line.width = 12
    monthly_line.height = 7
    ws.add_chart(monthly_line, "A7")

    # Monthly profit/loss bar chart
    monthly_pl = BarChart()
    monthly_pl.title = "Monthly Profit vs Loss"
    monthly_pl.style = 10
    monthly_pl.y_axis.title = "Amount"
    monthly_pl.x_axis.title = "Month"
    data = Reference(wb["Monthly Summary"], min_col=8, max_col=9, min_row=1, max_row=13)
    cats = Reference(wb["Monthly Summary"], min_col=3, min_row=2, max_row=13)
    monthly_pl.add_data(data, titles_from_data=True)
    monthly_pl.set_categories(cats)
    monthly_pl.width = 12
    monthly_pl.height = 7
    ws.add_chart(monthly_pl, "G7")

    # Yearly profit/loss bar chart
    yearly_pl = BarChart()
    yearly_pl.title = "Yearly Profit vs Loss"
    yearly_pl.style = 11
    yearly_pl.y_axis.title = "Amount"
    yearly_pl.x_axis.title = "Year"
    data = Reference(wb["Yearly Summary"], min_col=6, max_col=7, min_row=1, max_row=11)
    cats = Reference(wb["Yearly Summary"], min_col=1, min_row=2, max_row=11)
    yearly_pl.add_data(data, titles_from_data=True)
    yearly_pl.set_categories(cats)
    yearly_pl.width = 12
    yearly_pl.height = 7
    ws.add_chart(yearly_pl, "A22")

    # Yearly income/expense line chart
    yearly_ie = LineChart()
    yearly_ie.title = "Yearly Income vs Expenses"
    yearly_ie.style = 4
    yearly_ie.y_axis.title = "Amount"
    yearly_ie.x_axis.title = "Year"
    data = Reference(wb["Yearly Summary"], min_col=3, max_col=4, min_row=1, max_row=11)
    cats = Reference(wb["Yearly Summary"], min_col=1, min_row=2, max_row=11)
    yearly_ie.add_data(data, titles_from_data=True)
    yearly_ie.set_categories(cats)
    yearly_ie.width = 12
    yearly_ie.height = 7
    ws.add_chart(yearly_ie, "G22")

    # Route-wise top 5 tables (compatible formulas for wider Excel support)
    ws["R1"] = "Top 5 Profitable Routes"
    ws["S1"] = "Profit"
    ws["T1"] = "Top 5 Non-Profitable Routes"
    ws["U1"] = "Loss"

    months_array = (
        '{"January","February","March","April","May","June","July","August",'
        '"September","October","November","December"}'
    )
    ws["AA1"] = "Route Candidate"
    ws["AB1"] = "Route Profit"
    ws["AC1"] = "Route Loss"
    ws["AD1"] = "Profit Rank"
    ws["AE1"] = "Loss Rank"

    for row in range(2, 502):
        ws[f"AA{row}"] = "=IFERROR(INDEX(DailyRouteRecordTable[Route],ROW()-1),\"\")"
        ws[f"AB{row}"] = (
            f"=IF(AA{row}=\"\","
            "\"\",IF(COUNTIF($AA$2:AA"
            f"{row},AA{row})>1,0,"
            f"SUMPRODUCT((DailyRouteRecordTable[Route]=AA{row})"
            "*(IF($B$6=\"All\",1,IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=MATCH($B$6,"
            + months_array
            + ",0)))*N(DailyRouteRecordTable[Profit]))))"
        )
        ws[f"AC{row}"] = (
            f"=IF(AA{row}=\"\","
            "\"\",IF(COUNTIF($AA$2:AA"
            f"{row},AA{row})>1,0,"
            f"SUMPRODUCT((DailyRouteRecordTable[Route]=AA{row})"
            "*(IF($B$6=\"All\",1,IFERROR(--MID(DailyRouteRecordTable[Departure Date],4,2),0)=MATCH($B$6,"
            + months_array
            + ",0)))*N(DailyRouteRecordTable[Loss]))))"
        )
        ws[f"AD{row}"] = f"=IF(AB{row}<=0,\"\",RANK(AB{row},$AB$2:$AB$501,0)+COUNTIFS($AB$2:AB{row},AB{row})-1)"
        ws[f"AE{row}"] = f"=IF(AC{row}<=0,\"\",RANK(AC{row},$AC$2:$AC$501,0)+COUNTIFS($AC$2:AC{row},AC{row})-1)"

    ws["R2"] = "=IFERROR(INDEX($AA$2:$AA$501,MATCH(ROWS($R$2:R2),$AD$2:$AD$501,0)),\"\")"
    ws["S2"] = "=IF($R2=\"\",\"\",INDEX($AB$2:$AB$501,MATCH($R2,$AA$2:$AA$501,0)))"
    ws["T2"] = "=IFERROR(INDEX($AA$2:$AA$501,MATCH(ROWS($T$2:T2),$AE$2:$AE$501,0)),\"\")"
    ws["U2"] = "=IF($T2=\"\",\"\",INDEX($AC$2:$AC$501,MATCH($T2,$AA$2:$AA$501,0)))"
    for row in range(3, 7):
        ws[f"R{row}"] = f"=IFERROR(INDEX($AA$2:$AA$501,MATCH(ROWS($R$2:R{row}),$AD$2:$AD$501,0)),\"\")"
        ws[f"S{row}"] = f"=IF($R{row}=\"\",\"\",INDEX($AB$2:$AB$501,MATCH($R{row},$AA$2:$AA$501,0)))"
        ws[f"T{row}"] = f"=IFERROR(INDEX($AA$2:$AA$501,MATCH(ROWS($T$2:T{row}),$AE$2:$AE$501,0)),\"\")"
        ws[f"U{row}"] = f"=IF($T{row}=\"\",\"\",INDEX($AC$2:$AC$501,MATCH($T{row},$AA$2:$AA$501,0)))"

    for col in ["R", "S", "T", "U"]:
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="D9E1F2")
        ws[f"{col}1"].border = thin_border
    for row in range(2, 7):
        for col in ["R", "S", "T", "U"]:
            ws[f"{col}{row}"].border = thin_border
        ws[f"S{row}"].number_format = "#,##0.00"
        ws[f"U{row}"].number_format = "#,##0.00"

    # Route-wise charts (Top 5)
    route_profit_chart = BarChart()
    route_profit_chart.title = "Top 5 Profitable Routes (Selected Month)"
    route_profit_chart.style = 12
    route_profit_chart.y_axis.title = "Profit"
    route_profit_chart.x_axis.title = "Route"
    data = Reference(ws, min_col=19, max_col=19, min_row=1, max_row=6)  # S
    cats = Reference(ws, min_col=18, min_row=2, max_row=6)  # R
    route_profit_chart.add_data(data, titles_from_data=True)
    route_profit_chart.set_categories(cats)
    route_profit_chart.width = 12
    route_profit_chart.height = 7
    ws.add_chart(route_profit_chart, "A37")

    route_loss_chart = BarChart()
    route_loss_chart.title = "Top 5 Non-Profitable Routes (Selected Month)"
    route_loss_chart.style = 13
    route_loss_chart.y_axis.title = "Loss"
    route_loss_chart.x_axis.title = "Route"
    data = Reference(ws, min_col=21, max_col=21, min_row=1, max_row=6)  # U
    cats = Reference(ws, min_col=20, min_row=2, max_row=6)  # T
    route_loss_chart.add_data(data, titles_from_data=True)
    route_loss_chart.set_categories(cats)
    route_loss_chart.width = 12
    route_loss_chart.height = 7
    ws.add_chart(route_loss_chart, "G37")

    # Hide helper columns used for route calculations
    for col_letter in ["AA", "AB", "AC", "AD", "AE"]:
        ws.column_dimensions[col_letter].hidden = True


def main() -> None:
    workbook = build_workbook()
    workbook.save(OUTPUT_FILE)
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
